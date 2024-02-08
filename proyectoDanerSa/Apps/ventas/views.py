from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from Apps.inventario.models import Producto
from Apps.clientes.models import Cliente
from .models import Venta,DetalleVenta,Transaccion
from django.contrib import messages
from django.db import transaction
from decimal import Decimal, InvalidOperation
from django.shortcuts import get_object_or_404



#------Importaciones para reportes///
from django.utils import timezone 
from openpyxl.styles import Font, PatternFill, Border, Side
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from django.db.models import Sum






@login_required(login_url="/accounts/login/")
def ListaVentasView(request):
    ventas = Venta.objects.all()
    context = {
        "ventas": ventas,
    }
    return render(request, "ventas/ventas.html", context=context)


@login_required(login_url="/accounts/login/")
@transaction.atomic
def VistaAgregarVentas(request):
    if request.method == 'POST':
        try:
            cliente_id = request.POST.get('cliente')
            anticipo = Decimal(request.POST.get('anticipo', 0))
            total = Decimal(request.POST.get('total', 0))
            comentario = request.POST.get('comentario', '')
            saldo = Decimal(request.POST.get('saldo', 0))
            estadoCuenta = saldo != Decimal(0)
            

            with transaction.atomic():
                cliente = Cliente.objects.get(pk=cliente_id)
                cliente.saldo += saldo
                cliente.save()

                venta = Venta.objects.create(
                    id_clientes=cliente,
                    anticipo=anticipo,
                    total=total,
                    comentario=comentario,
                    estado_cuenta=estadoCuenta,
                    user = request.user,
                )

                crear_detalles_venta(request, venta)

                messages.success(request, '¡Venta creada con éxito!', extra_tags="success")
        except Cliente.DoesNotExist:
            messages.error(request, 'El cliente no existe', extra_tags="danger")
        except InvalidOperation as e:
            messages.error(request, f'Error en operación decimal: {str(e)}', extra_tags="danger")
        except Exception as e:
            messages.error(request, f'Error al procesar la venta: {str(e)}', extra_tags="danger")
        return redirect('Apps.ventas:lista_ventas')

    productos = Producto.objects.all()
    clientes = Cliente.objects.all()
    context = {
        "productos": productos,
        "clientes": clientes,
    }
    return render(request, "ventas/agregar_ventas.html", context=context)

@login_required(login_url="/accounts/login/")
def abonar_transaccion(request, id_venta):
    venta = get_object_or_404(Venta, pk=id_venta)
    abono = Decimal(request.POST.get('abono',0))
    cliente = venta.id_clientes
    try:
        with transaction.atomic():
            saldo_venta = Decimal('0.0') if venta.getSaldo() is None else venta.getSaldo()
            saldo_cliente = Decimal('0.0') if cliente.saldo is None else cliente.saldo

            cliente.saldo = saldo_cliente - abono
            cliente.save()

            if saldo_venta - abono == 0:
                venta.estado_cuenta = False
                venta.save()
                Transaccion.objects.create(
                id_cliente=cliente,
                id_venta=venta,
                abono=abono,
                saldo_venta=saldo_venta - abono,
                saldo_cliente=saldo_cliente - abono
                )
            else:
                ultima_transaccion = Transaccion.objects.filter(id_venta=id_venta).order_by('-id').first()
                if ultima_transaccion:
                    restante = ultima_transaccion.saldo_venta - abono
                    if restante == 0:
                        venta.estado_cuenta = False
                        venta.save()
                        Transaccion.objects.create(
                            id_cliente=cliente,
                            id_venta=venta,
                            abono=abono,
                            saldo_venta=ultima_transaccion.saldo_venta - abono,
                            saldo_cliente=saldo_cliente - abono
                        )
                    else:
                            Transaccion.objects.create(
                            id_cliente=cliente,
                            id_venta=venta,
                            abono=abono,
                            saldo_venta=ultima_transaccion.saldo_venta - abono,
                            saldo_cliente=saldo_cliente - abono
                        )
                else: 
                        Transaccion.objects.create(
                            id_cliente=cliente,
                            id_venta=venta,
                            abono=abono,
                            saldo_venta=saldo_venta - abono,
                            saldo_cliente=saldo_cliente - abono
                        )
            messages.success(request, '¡El abono fue un éxito!', extra_tags="success")

    except Exception as e:
        messages.error(request, f'Error al procesar el abono: {str(e)}', extra_tags="danger")

    return redirect('Apps.clientes:detalles_clientes',id_cliente=cliente.id )

@login_required(login_url="/accounts/login/")
def crear_detalles_venta(request, venta):
    productos_ids = request.POST.getlist('productos_ids[]')
    productos_cantidades = request.POST.getlist('productos_cantidades[]')
    productos_precios = request.POST.getlist('productos_precios[]')

    for productos, cantidades, precios in zip(productos_ids, productos_cantidades, productos_precios):
        lista_productos = productos.split(',')
        lista_cantidades = cantidades.split(',')
        lista_precios = precios.split(',')

        for producto_id, cantidad, precio in zip(lista_productos, lista_cantidades, lista_precios):
            producto = Producto.objects.get(idproducto=producto_id)
            cantidad_vendida = int(cantidad)

            if producto.existencia >= cantidad_vendida:
                producto.existencia -= cantidad_vendida
                producto.save()
            else:
                raise Exception('No hay suficiente existencia para realizar la venta.')
            ganancia = Decimal(precio) - producto.precio_compra 

            DetalleVenta.objects.create(
                id_venta=venta,
                id_producto=producto,
                cantidad=cantidad,
                precio=precio,
                ganancia=ganancia
            )



            

#------------------------------------Reportes para ventas --------------------------------------
            

#---lleva a pagina de reportes     
@login_required(login_url="/accounts/login/")
def VentasView(request):

    return render(request, "ventas/Vreport.html")                   



@login_required(login_url="/accounts/login/")
def generar_reporte_excel(request, start_date, end_date):
    start_date = timezone.datetime.strptime(start_date, '%Y-%m-%d')
    end_date = timezone.datetime.strptime(end_date, '%Y-%m-%d')

    ventas = Venta.objects.filter(fecha__range=[start_date, end_date])

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=reporte_ventas.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Ventas"

    headers = ['Usuario', 'Fecha', 'Cliente', 'Total', 'Anticipo', 'Saldo Restante', 'Productos Vendidos', 'Cantidad Vendida', 'Precio de Compra', 'Precio de Venta', 'Ganancia']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=2, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    for row_num, venta in enumerate(ventas, 3):
        detalles_venta = DetalleVenta.objects.filter(id_venta=venta)

        user_info = f"{venta.user.username}"
        worksheet.cell(row=row_num, column=1, value=user_info).alignment = Alignment(horizontal='left', wrap_text=True)

        worksheet.cell(row=row_num, column=2, value=venta.fecha).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=3, value=venta.id_clientes.nombre).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=4, value=venta.total).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=5, value=venta.anticipo).alignment = Alignment(horizontal='center')

        productos_vendidos = ', '.join([detalle.id_producto.nombre for detalle in detalles_venta])
        worksheet.cell(row=row_num, column=7, value=productos_vendidos).alignment = Alignment(horizontal='center')

        total_cantidad = detalles_venta.aggregate(Sum('cantidad'))['cantidad__sum']
        total_ganancia = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']

        worksheet.cell(row=row_num, column=8, value=total_cantidad).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=11, value=total_ganancia).alignment = Alignment(horizontal='center')

        # Nuevas líneas para obtener el precio de compra y precio de venta
        precio_venta = detalles_venta.aggregate(Sum('precio'))['precio__sum']
        ganancia_total = detalles_venta.aggregate(Sum('ganancia'))['ganancia__sum']
        precio_compra = precio_venta - ganancia_total

        worksheet.cell(row=row_num, column=9, value=precio_compra).alignment = Alignment(horizontal='center')
        worksheet.cell(row=row_num, column=10, value=precio_venta).alignment = Alignment(horizontal='center')

        ultima_transaccion = Transaccion.objects.filter(id_venta=venta).order_by('-fecha').first()
        saldo_restante = ultima_transaccion.saldo_venta if ultima_transaccion else 0

        worksheet.cell(row=row_num, column=6, value=saldo_restante).alignment = Alignment(horizontal='center')

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)

    return response





@login_required(login_url="/accounts/login/")
def generar_reporte_cliente_excel(request, id_cliente):
    cliente = get_object_or_404(Cliente, pk=id_cliente)
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=reporte_cliente_{cliente.nombre}.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Reporte de {cliente.nombre}"

    title_font = Font(size=14, color='FFFFFF', bold=True)
    title_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    title_border = Border(left=Side(border_style='thin', color='000000'),
                          right=Side(border_style='thin', color='000000'),
                          top=Side(border_style='thin', color='000000'),
                          bottom=Side(border_style='thin', color='000000'))

    headers = ['Cliente', 'Fecha', 'Abono', 'Saldo Restante', 'Producto Vendido', 'Venta Total']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=1, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    transacciones_por_venta = Transaccion.objects.filter(id_cliente=cliente, fecha__range=[start_date, end_date]).values('id_venta').distinct()
    row_num = 2

    for transaccion_venta in transacciones_por_venta:
        id_venta = transaccion_venta['id_venta']
        transacciones_venta_actual = Transaccion.objects.filter(id_cliente=cliente, id_venta=id_venta, fecha__range=[start_date, end_date])

        for i, transaccion in enumerate(transacciones_venta_actual):
            worksheet.cell(row=row_num, column=1, value=transaccion.id_venta.id_clientes.nombre).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=2, value=transaccion.fecha).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=3, value=transaccion.abono).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=4, value=transaccion.saldo_venta).alignment = Alignment(horizontal='center')

            productos_vendidos = ', '.join(transaccion.id_venta.detalleventa_set.all().values_list('id_producto__nombre', flat=True)) 
            worksheet.cell(row=row_num, column=5, value=productos_vendidos).alignment = Alignment(horizontal='center')

            if i == 0:
                total_venta_actual = Venta.objects.filter(id=id_venta, fecha__range=[start_date, end_date]).aggregate(Sum('total'))['total__sum'] or Decimal(0)
                total_venta_cell = worksheet.cell(row=row_num, column=6, value=total_venta_actual).alignment = Alignment(horizontal='center')

            row_num += 1

        row_num += 2

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)
    return response





@login_required(login_url="/accounts/login/")
def generar_reporte_abonos_clientes_excel(request, start_date, end_date):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename=reporte_abonos_clientes.xlsx'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Reporte de Abonos Clientes"

    title_font = Font(size=14, color='FFFFFF', bold=True)
    title_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    title_border = Border(left=Side(border_style='thin', color='000000'),
                          right=Side(border_style='thin', color='000000'),
                          top=Side(border_style='thin', color='000000'),
                          bottom=Side(border_style='thin', color='000000'))

    headers = ['Cliente', 'Fecha', 'Abono', 'Saldo Restante', 'Producto Vendido', 'Venta Total']

    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = worksheet.cell(row=1, column=col_num, value=header)

        cell.fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
        cell.border = Border(left=Side(border_style='thin', color='000000'),
                             right=Side(border_style='thin', color='000000'),
                             top=Side(border_style='thin', color='000000'),
                             bottom=Side(border_style='thin', color='000000'))
        cell.alignment = Alignment(horizontal='center')

    transacciones_por_venta = Transaccion.objects.filter(fecha__range=[start_date, end_date]).values('id_venta').distinct()
    row_num = 2

    for transaccion_venta in transacciones_por_venta:
        id_venta = transaccion_venta['id_venta']
        transacciones_venta_actual = Transaccion.objects.filter(id_venta=id_venta, fecha__range=[start_date, end_date])

        for i, transaccion in enumerate(transacciones_venta_actual):
            worksheet.cell(row=row_num, column=1, value=transaccion.id_venta.id_clientes.nombre).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=2, value=transaccion.fecha).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=3, value=transaccion.abono).alignment = Alignment(horizontal='center')
            worksheet.cell(row=row_num, column=4, value=transaccion.saldo_venta).alignment = Alignment(horizontal='center')

            productos_vendidos = ', '.join(transaccion.id_venta.detalleventa_set.all().values_list('id_producto__nombre', flat=True)) 
            worksheet.cell(row=row_num, column=5, value=productos_vendidos).alignment = Alignment(horizontal='center')

            if i == 0:
                total_venta_actual = Venta.objects.filter(id=id_venta, fecha__range=[start_date, end_date]).aggregate(Sum('total'))['total__sum'] or Decimal(0)
                total_venta_cell = worksheet.cell(row=row_num, column=6, value=total_venta_actual).alignment = Alignment(horizontal='center')

            row_num += 1

        row_num += 2

    for col_num in range(1, worksheet.max_column + 1):
        max_length = 0
        column = get_column_letter(col_num)
        for cell in worksheet[column]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column].width = adjusted_width

    workbook.save(response)
    return response